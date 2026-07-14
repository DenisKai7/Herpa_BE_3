"""Report generation v5: Rich tables, CSV, JSON, Markdown, Excel, HTML, charts.

GraphRAG-optimized: Only Answer Relevancy as LLM metric.
Removed: Faithfulness, Context Precision, Context Recall, Context Relevancy, Hallucination.
"""

import csv
import json
import logging
from pathlib import Path
from typing import Any

from evaluation.metrics.latency import format_latency_ms

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).parent / "results"

# Removed metric keys for backward compatibility
_REMOVED_METRICS = {"faithfulness", "contextual_precision", "contextual_recall", "contextual_relevancy", "hallucination_score"}


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Safely convert value to float."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val)
        except (ValueError, TypeError):
            return default
    return default


def _ensure_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ─── Rich Terminal Report ────────────────────────────────────────────────────

def print_terminal_report(eval_results: dict[str, Any]):
    """Print formatted report using Rich tables."""
    try:
        from rich.console import Console
        from rich.table import Table
        from rich.panel import Panel
        from rich import box
        _print_rich(eval_results)
    except ImportError:
        _print_plain(eval_results)


def _print_rich(eval_results: dict[str, Any]):
    """Rich-formatted terminal output with comprehensive dashboards."""
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich import box

    console = Console()
    agg = eval_results.get("aggregated_metrics", {})
    total = eval_results.get("total_queries", 0)
    cached = eval_results.get("cached_count", 0)
    success = eval_results.get("success_count", 0)
    skipped = eval_results.get("skip_count", 0)
    failed = eval_results.get("fail_count", 0)
    mode = eval_results.get("mode", "full")
    profiling = eval_results.get("profiling", {})
    llm_calls = eval_results.get("llm_calls", 0)
    neo4j_calls = eval_results.get("neo4j_calls", 0)
    analytics = eval_results.get("analytics", {})
    health = eval_results.get("health_check", {})
    cache_stats = eval_results.get("cache_stats", {})

    retrieval = agg.get("retrieval", {})
    judge = agg.get("judge", {})
    citation = agg.get("citation", {})
    graph = agg.get("graph", {})
    latency = agg.get("latency", {})
    overall = agg.get("overall_score", {})
    judge_count = agg.get("judge_query_count", 0)

    # Header
    status_line = f"[green]{success}[/] passed"
    if skipped:
        status_line += f" | [yellow]{skipped}[/] skipped"
    if failed:
        status_line += f" | [red]{failed}[/] failed"

    console.print()
    console.print(Panel.fit(
        f"[bold cyan]HERPA GraphRAG Evaluation[/]\n"
        f"Dataset: {total} queries | Mode: {mode} | Cached: {cached}\n"
        f"{status_line}\n"
        f"LLM Calls: {llm_calls} | Neo4j Calls: {neo4j_calls} | Judge Success: {judge_count}",
        border_style="cyan"
    ))

    # ── Diagnostic Table ──
    judge_analytics = analytics.get("judge", {})
    cache_hits = cache_stats.get("hits", {})
    cache_misses = cache_stats.get("misses", {})
    total_cache = sum(cache_hits.values()) + sum(cache_misses.values())
    cache_hit_pct = (sum(cache_hits.values()) / total_cache * 100) if total_cache > 0 else 0

    diag = Table(title="Evaluation Diagnostics", box=box.HEAVY_EDGE, show_lines=False)
    diag.add_column("Key", style="bold")
    diag.add_column("Value", justify="right")
    diag.add_row("Judge Server", health.get("llm_server", "N/A"))
    diag.add_row("Neo4j", health.get("neo4j", "N/A"))
    diag.add_row("Retriever", health.get("retriever", "N/A"))
    diag.add_row("Judge Model", str(judge_analytics.get("judge_model", "N/A")))
    diag.add_row("Prompt Tokens (total)", str(judge_analytics.get("prompt_tokens", 0)))
    diag.add_row("Avg Context Chars", f"{sum(r.get('context_size', 0) for r in eval_results.get('per_query_results', [])) / max(1, total):.0f}")
    diag.add_row("Retry Count", str(judge_analytics.get("retry_count", 0)))
    diag.add_row("Cache Hit", f"{cache_hit_pct:.0f}%")
    diag.add_row("Pipeline Errors", str(failed))
    diag.add_row("Judge Errors", str(judge_analytics.get("failure_count", 0)))
    diag.add_row("Judge Timeouts", str(judge_analytics.get("timeout_count", 0)))
    console.print(diag)

    # Retrieval table
    t = Table(title="Retrieval Metrics", box=box.ROUNDED, show_lines=False)
    t.add_column("Metric", style="bold")
    t.add_column("Value", justify="right")
    for k in ["precision@1", "precision@3", "precision@5"]:
        t.add_row(k, f"{retrieval.get(k, 0):.4f}")
    for k in ["recall@1", "recall@3", "recall@5"]:
        t.add_row(k, f"{retrieval.get(k, 0):.4f}")
    for k in ["hit_rate@1", "hit_rate@3", "hit_rate@5"]:
        t.add_row(k, f"{retrieval.get(k, 0) * 100:.1f}%")
    t.add_row("MRR", f"{retrieval.get('mrr', 0):.4f}")
    t.add_row("nDCG", f"{retrieval.get('ndcg', 0):.4f}")
    console.print(t)

    # LLM Evaluation table (only Answer Relevancy)
    t2 = Table(title="LLM Evaluation", box=box.ROUNDED, show_lines=False)
    t2.add_column("Metric", style="bold")
    t2.add_column("Score", justify="right")
    ans_rel = _safe_float(judge.get("answer_relevancy"))
    style = "green" if ans_rel >= 0.8 else "yellow" if ans_rel >= 0.6 else "red"
    t2.add_row("Answer Relevancy", f"[{style}]{ans_rel:.4f}[/]")
    console.print(t2)

    # Citation & Graph table
    t3 = Table(title="Citation & Graph", box=box.ROUNDED, show_lines=False)
    t3.add_column("Metric", style="bold")
    t3.add_column("Value", justify="right")
    t3.add_row("Citation Accuracy", f"{_safe_float(citation.get('accuracy')) * 100:.1f}%")
    t3.add_row("Node Precision", f"{_safe_float(graph.get('node_precision')) * 100:.1f}%")
    t3.add_row("Node Recall", f"{_safe_float(graph.get('node_recall')) * 100:.1f}%")
    t3.add_row("Relation Precision", f"{_safe_float(graph.get('relationship_precision')) * 100:.1f}%")
    t3.add_row("Relation Recall", f"{_safe_float(graph.get('relationship_recall')) * 100:.1f}%")
    t3.add_row("Graph Accuracy", f"{_safe_float(graph.get('overall_accuracy')) * 100:.1f}%")
    console.print(t3)

    # Latency table
    t4 = Table(title="Latency", box=box.ROUNDED, show_lines=False)
    t4.add_column("Stage", style="bold")
    t4.add_column("Avg", justify="right")
    t4.add_column("Median", justify="right")
    t4.add_column("P95", justify="right")
    t4.add_column("P99", justify="right")
    for stage in ["embedding", "neo4j", "retrieval", "llm", "judge", "total"]:
        stats = latency.get(stage, {})
        if not stats or stats.get("avg", 0) == 0:
            continue
        t4.add_row(
            stage.capitalize(),
            format_latency_ms(stats.get("avg", 0)),
            format_latency_ms(stats.get("median", 0)),
            format_latency_ms(stats.get("p95", 0)),
            format_latency_ms(stats.get("p99", 0)),
        )
    console.print(t4)

    # Overall Evaluation Dashboard
    console.print()
    t_dash = Table(title="OVERALL EVALUATION", box=box.DOUBLE_EDGE, show_lines=True)
    t_dash.add_column("Component", style="bold")
    t_dash.add_column("Score", justify="right")
    t_dash.add_column("Weight", justify="right")
    component_scores = overall.get("component_scores", {})
    component_weights = overall.get("weights", {})
    component_labels = {
        "retrieval": "Retrieval",
        "answer_relevancy": "Generation",
        "graph": "Graph",
        "citation": "Citation",
        "latency": "Performance",
    }
    for key, label in component_labels.items():
        val = _safe_float(component_scores.get(key))
        w = component_weights.get(key, 0)
        style = "green" if val >= 80 else "yellow" if val >= 60 else "red"
        t_dash.add_row(label, f"[{style}]{val:.1f}[/]", f"{w:.0%}")
    score = _safe_float(overall.get("score"))
    grade = overall.get("grade", "N/A")
    score_style = "bold green" if score >= 80 else "bold yellow" if score >= 60 else "bold red"
    t_dash.add_row("[bold]Overall Score[/]", f"[{score_style}]{score:.2f}[/]", "")
    t_dash.add_row("[bold]Grade[/]", f"[bold]{grade}[/]", "")
    console.print(t_dash)

    # Sub-scores detail
    sub_scores = overall.get("sub_scores", {})
    if sub_scores:
        console.print()
        t_sub = Table(title="Score Breakdown Detail", box=box.SIMPLE_HEAVY, show_lines=False)
        t_sub.add_column("Component", style="bold")
        t_sub.add_column("Sub-Metric", style="dim")
        t_sub.add_column("Value", justify="right")

        ret_sub = sub_scores.get("retrieval", {})
        for k, v in ret_sub.items():
            t_sub.add_row("Retrieval", k, f"{v:.4f}")

        graph_sub = sub_scores.get("graph", {})
        for k, v in graph_sub.items():
            t_sub.add_row("Graph", k.replace("_", " ").title(), f"{v:.4f}")

        perf_sub = sub_scores.get("performance", {})
        if perf_sub:
            t_sub.add_row("Performance", "Neo4j weight", f"{perf_sub.get('neo4j_weight', 0):.0%}")
            t_sub.add_row("Performance", "Retrieval weight", f"{perf_sub.get('retrieval_weight', 0):.0%}")
            t_sub.add_row("Performance", "Judge weight", f"{perf_sub.get('judge_weight', 0):.0%}")

        console.print(t_sub)

    # Judge Summary
    if judge_analytics:
        console.print()
        t_judge = Table(title="Judge Summary", box=box.ROUNDED, show_lines=False)
        t_judge.add_column("Metric", style="bold")
        t_judge.add_column("Value", justify="right")
        t_judge.add_row("Judge Model", str(judge_analytics.get("judge_model", "N/A")))
        t_judge.add_row("Prompt Version", str(judge_analytics.get("judge_prompt_version", "N/A")))
        t_judge.add_row("Total Query", str(judge_analytics.get("success_count", 0) + judge_analytics.get("failure_count", 0)))
        t_judge.add_row("Judge Success", str(judge_analytics.get("success_count", 0)))
        t_judge.add_row("Judge Timeout", str(judge_analytics.get("timeout_count", 0)))
        t_judge.add_row("Judge Retry", str(judge_analytics.get("retry_count", 0)))
        t_judge.add_row("Judge Failed", str(judge_analytics.get("failure_count", 0)))
        t_judge.add_row("Average Judge Time", f"{judge_analytics.get('average_judge_time_ms', 0):.0f} ms")
        t_judge.add_row("Prompt Tokens", str(judge_analytics.get("prompt_tokens", 0)))
        t_judge.add_row("Completion Tokens", str(judge_analytics.get("completion_tokens", 0)))
        t_judge.add_row("Total Tokens", str(judge_analytics.get("total_tokens", 0)))
        console.print(t_judge)

    # Retrieval Summary
    retrieval_analytics = analytics.get("retrieval", {})
    console.print()
    t_ret_sum = Table(title="Retrieval Summary", box=box.ROUNDED, show_lines=False)
    t_ret_sum.add_column("Metric", style="bold")
    t_ret_sum.add_column("Value", justify="right")
    t_ret_sum.add_row("Average Retrieved Nodes", f"{retrieval_analytics.get('average_retrieved_nodes', 0):.1f}")
    t_ret_sum.add_row("Average Retrieved Relations", f"{retrieval_analytics.get('average_retrieved_relations', 0):.1f}")
    t_ret_sum.add_row("Average Graph Depth", f"{retrieval_analytics.get('average_graph_depth', 0):.1f}")
    t_ret_sum.add_row("Average Graph Breadth", f"{retrieval_analytics.get('average_graph_breadth', 0):.1f}")
    console.print(t_ret_sum)

    # Neo4j Analytics
    neo4j_analytics = analytics.get("neo4j", {})
    if neo4j_analytics:
        console.print()
        t_neo = Table(title="Neo4j Analytics", box=box.ROUNDED, show_lines=False)
        t_neo.add_column("Metric", style="bold")
        t_neo.add_column("Value", justify="right")
        t_neo.add_row("Neo4j Calls", str(neo4j_analytics.get("neo4j_calls", 0)))
        t_neo.add_row("Average Query Time", f"{neo4j_analytics.get('average_query_time_ms', 0):.0f} ms")
        t_neo.add_row("Median Query Time", f"{neo4j_analytics.get('median_query_time_ms', 0):.0f} ms")
        t_neo.add_row("Fastest Query", f"{neo4j_analytics.get('fastest_query_ms', 0):.0f} ms")
        t_neo.add_row("Slowest Query", f"{neo4j_analytics.get('slowest_query_ms', 0):.0f} ms")
        t_neo.add_row("Total Nodes Retrieved", str(neo4j_analytics.get("total_nodes_retrieved", 0)))
        t_neo.add_row("Total Relations Retrieved", str(neo4j_analytics.get("total_relations_retrieved", 0)))
        t_neo.add_row("Average Nodes/Query", f"{neo4j_analytics.get('average_nodes_per_query', 0):.1f}")
        t_neo.add_row("Average Relations/Query", f"{neo4j_analytics.get('average_relations_per_query', 0):.1f}")
        console.print(t_neo)

    # Failure Analysis
    failure_analysis = analytics.get("failure_analysis", {})
    if failure_analysis and failure_analysis.get("total_failures", 0) > 0:
        console.print()
        t_fail = Table(title=f"Failure Analysis ({failure_analysis['total_failures']} failures)", box=box.ROUNDED, show_lines=True)
        t_fail.add_column("Category", style="bold")
        t_fail.add_column("Count", justify="right")
        for cat, count in failure_analysis.get("categories", {}).items():
            t_fail.add_row(cat.replace("_", " ").title(), str(count))
        console.print(t_fail)

    # Profiling
    total_time = eval_results.get("total_profiling_s", 0)
    if profiling:
        t5 = Table(title="Profiling", box=box.ROUNDED, show_lines=False)
        t5.add_column("Stage", style="bold")
        t5.add_column("Time", justify="right")
        t5.add_column("%", justify="right")
        total_time = total_time or sum(profiling.values())
        for stage, elapsed in sorted(profiling.items(), key=lambda x: -x[1]):
            pct = (elapsed / total_time * 100) if total_time > 0 else 0
            t5.add_row(stage, f"{elapsed:.1f}s", f"{pct:.0f}%")
        t5.add_row("[bold]TOTAL[/]", f"[bold]{total_time:.1f}s[/]", "100%")
        console.print(t5)

    # Performance stats
    queries_per_sec = total / total_time if total_time > 0 else 0
    avg_per_query = total_time / total if total > 0 else 0
    cache_ratio = cached / total * 100 if total > 0 else 0
    console.print(f"\n  Queries/sec: {queries_per_sec:.1f} | Avg/query: {avg_per_query:.1f}s | "
                  f"LLM calls: {llm_calls} | Neo4j calls: {neo4j_calls} | Cache hit: {cache_ratio:.0f}%")
    console.print()


def _print_plain(eval_results: dict[str, Any]):
    """Fallback plain text output."""
    agg = eval_results.get("aggregated_metrics", {})
    total = eval_results.get("total_queries", 0)
    overall = agg.get("overall_score", {})
    retrieval = agg.get("retrieval", {})
    judge = agg.get("judge", {})
    citation = agg.get("citation", {})
    graph = agg.get("graph", {})
    profiling = eval_results.get("profiling", {})

    w = 30
    print()
    print("=" * (w + 18))
    print(f"{'HERPA GraphRAG Evaluation':^{w + 16}}")
    print(f"{'Dataset: ' + str(total) + ' queries':^{w + 16}}")
    print("=" * (w + 18))

    print("\n  Retrieval Metrics")
    for k in ["precision@1", "precision@3", "precision@5", "recall@1", "recall@3", "recall@5", "mrr", "ndcg"]:
        print(f"    {k:<{w}} {retrieval.get(k, 0):.4f}")
    for k in ["hit_rate@1", "hit_rate@3", "hit_rate@5"]:
        print(f"    {k:<{w}} {retrieval.get(k, 0) * 100:.1f}%")

    print(f"\n  LLM Evaluation")
    print(f"    {'Answer Relevancy':<{w}} {_safe_float(judge.get('answer_relevancy')):.4f}")

    print(f"\n  Citation & Graph")
    print(f"    {'Citation Accuracy':<{w}} {_safe_float(citation.get('accuracy')) * 100:.1f}%")
    print(f"    {'Node Precision':<{w}} {_safe_float(graph.get('node_precision')) * 100:.1f}%")
    print(f"    {'Node Recall':<{w}} {_safe_float(graph.get('node_recall')) * 100:.1f}%")
    print(f"    {'Relation Precision':<{w}} {_safe_float(graph.get('relationship_precision')) * 100:.1f}%")
    print(f"    {'Relation Recall':<{w}} {_safe_float(graph.get('relationship_recall')) * 100:.1f}%")
    print(f"    {'Graph Accuracy':<{w}} {_safe_float(graph.get('overall_accuracy')) * 100:.1f}%")

    # Overall breakdown
    component_scores = overall.get("component_scores", {})
    component_weights = overall.get("weights", {})
    print(f"\n  Overall Score Breakdown")
    for key, label in [("retrieval","Retrieval"),("answer_relevancy","Generation"),("graph","Graph"),("citation","Citation"),("latency","Performance")]:
        val = _safe_float(component_scores.get(key))
        w = component_weights.get(key, 0)
        print(f"    {label:<{w}} {val:.1f}  (weight: {w:.0%})")

    print(f"\n  OVERALL: {overall.get('score', 0):.2f}/100  Grade: {overall.get('grade', 'N/A')}")

    if profiling:
        print("\n  Profiling:")
        for stage, elapsed in sorted(profiling.items(), key=lambda x: -x[1]):
            print(f"    {stage:<{w}} {elapsed:.1f}s")
    print()


# ─── Export Functions ────────────────────────────────────────────────────────

def save_json(eval_results: dict[str, Any]) -> str:
    _ensure_dir()
    path = OUTPUT_DIR / "evaluation_report.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(eval_results, f, ensure_ascii=False, indent=2, default=str)
    return str(path)


def save_summary_json(eval_results: dict[str, Any]) -> str:
    _ensure_dir()
    path = OUTPUT_DIR / "summary.json"
    summary = {
        "total_queries": eval_results.get("total_queries", 0),
        "cached_count": eval_results.get("cached_count", 0),
        "mode": eval_results.get("mode", "full"),
        "aggregated_metrics": eval_results.get("aggregated_metrics", {}),
        "analytics": eval_results.get("analytics", {}),
        "profiling": eval_results.get("profiling", {}),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
    return str(path)


def save_csv(eval_results: dict[str, Any]) -> str:
    _ensure_dir()
    path = OUTPUT_DIR / "evaluation_report.csv"
    per_query = eval_results.get("per_query_results", [])
    if not per_query:
        return str(path)

    rows = []
    for r in per_query:
        row = {
            "id": r.get("id"), "query": r.get("query"), "category": r.get("category"),
            "answer": r.get("answer", "")[:500], "ground_truth": r.get("ground_truth", "")[:500],
        }
        ret = r.get("retrieval_metrics", {})
        for k, v in ret.items():
            row[f"retrieval_{k}"] = v
        row["answer_relevancy"] = _safe_float(r.get("answer_relevancy"))
        row["citation_accuracy"] = r.get("citation_metrics", {}).get("accuracy", 0)
        graph = r.get("graph_accuracy", {})
        for k in ["node_precision", "node_recall", "relationship_precision", "relationship_recall", "overall_accuracy"]:
            row[f"graph_{k}"] = graph.get(k, 0)
        lat = r.get("latency", {})
        for k in ["retrieval_ms", "llm_ms", "judge_ms", "total_ms"]:
            row[f"latency_{k}"] = lat.get(k, 0)
        row["context_tokens"] = r.get("context_tokens", 0)
        row["status"] = r.get("status", "")
        rows.append(row)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    return str(path)


def save_markdown(eval_results: dict[str, Any]) -> str:
    _ensure_dir()
    path = OUTPUT_DIR / "evaluation_report.md"

    agg = eval_results.get("aggregated_metrics", {})
    total = eval_results.get("total_queries", 0)
    retrieval = agg.get("retrieval", {})
    judge = agg.get("judge", {})
    citation = agg.get("citation", {})
    graph = agg.get("graph", {})
    latency = agg.get("latency", {})
    overall = agg.get("overall_score", {})
    profiling = eval_results.get("profiling", {})
    analytics = eval_results.get("analytics", {})

    md = []
    md.append("# Evaluasi Sistem GraphRAG & Agentic AI HERPA\n")
    md.append(f"**Jumlah Query:** {total}\n")
    md.append(f"**Mode:** {eval_results.get('mode', 'full')}\n")
    md.append(f"**Overall Score:** {overall.get('score', 0):.2f} / 100\n")
    md.append(f"**Grade:** {overall.get('grade', 'N/A')}\n")

    md.append("\n## Retrieval Evaluation\n")
    md.append("| Metric | Value |")
    md.append("|--------|-------|")
    for k in ["precision@1", "precision@3", "precision@5", "recall@1", "recall@3", "recall@5"]:
        md.append(f"| {k} | {retrieval.get(k, 0):.4f} |")
    for k in ["hit_rate@1", "hit_rate@3", "hit_rate@5"]:
        md.append(f"| {k} | {retrieval.get(k, 0) * 100:.1f}% |")
    md.append(f"| MRR | {retrieval.get('mrr', 0):.4f} |")
    md.append(f"| nDCG | {retrieval.get('ndcg', 0):.4f} |")

    md.append("\n## Generation (LLM Evaluation)\n")
    md.append("| Metric | Score |")
    md.append("|--------|-------|")
    md.append(f"| Answer Relevancy | {_safe_float(judge.get('answer_relevancy')):.4f} |")

    md.append("\n## Citation & Graph\n")
    md.append("| Metric | Value |")
    md.append("|--------|-------|")
    md.append(f"| Citation Accuracy | {_safe_float(citation.get('accuracy')) * 100:.1f}% |")
    md.append(f"| Node Precision | {_safe_float(graph.get('node_precision')) * 100:.1f}% |")
    md.append(f"| Node Recall | {_safe_float(graph.get('node_recall')) * 100:.1f}% |")
    md.append(f"| Relationship Precision | {_safe_float(graph.get('relationship_precision')) * 100:.1f}% |")
    md.append(f"| Relationship Recall | {_safe_float(graph.get('relationship_recall')) * 100:.1f}% |")
    md.append(f"| Graph Accuracy | {_safe_float(graph.get('overall_accuracy')) * 100:.1f}% |")

    md.append("\n## Latency\n")
    md.append("| Stage | Avg | Median | P95 | P99 |")
    md.append("|-------|-----|--------|-----|-----|")
    for stage in ["embedding", "neo4j", "retrieval", "llm", "judge", "total"]:
        stats = latency.get(stage, {})
        if not stats or stats.get("avg", 0) == 0:
            continue
        md.append(f"| {stage.capitalize()} | {format_latency_ms(stats.get('avg', 0))} | {format_latency_ms(stats.get('median', 0))} | {format_latency_ms(stats.get('p95', 0))} | {format_latency_ms(stats.get('p99', 0))} |")

    md.append("\n## Overall Score Breakdown\n")
    md.append("| Component | Score | Weight |")
    md.append("|-----------|-------|--------|")
    component_scores = overall.get("component_scores", {})
    weights = overall.get("weights", {})
    for key, score in component_scores.items():
        w = weights.get(key, 0)
        label = key.replace("_", " ").title()
        md.append(f"| {label} | {score:.1f} | {w:.0%} |")

    # Judge Summary
    judge_analytics = analytics.get("judge", {})
    if judge_analytics:
        md.append("\n## Judge Summary\n")
        md.append("| Metric | Value |")
        md.append("|--------|-------|")
        md.append(f"| Judge Model | {judge_analytics.get('judge_model', 'N/A')} |")
        md.append(f"| Total Query | {judge_analytics.get('success_count', 0) + judge_analytics.get('failure_count', 0)} |")
        md.append(f"| Judge Success | {judge_analytics.get('success_count', 0)} |")
        md.append(f"| Judge Timeout | {judge_analytics.get('timeout_count', 0)} |")
        md.append(f"| Average Judge Time | {judge_analytics.get('average_judge_time_ms', 0):.0f} ms |")
        md.append(f"| Total Tokens | {judge_analytics.get('total_tokens', 0)} |")

    # Failure Analysis
    failure_analysis = analytics.get("failure_analysis", {})
    if failure_analysis and failure_analysis.get("total_failures", 0) > 0:
        md.append("\n## Failure Analysis\n")
        md.append("| Category | Count |")
        md.append("|----------|-------|")
        for cat, count in failure_analysis.get("categories", {}).items():
            md.append(f"| {cat.replace('_', ' ').title()} | {count} |")

    if profiling:
        md.append("\n## Profiling\n")
        md.append("| Stage | Time | % |")
        md.append("|-------|------|---|")
        total_time = eval_results.get("total_profiling_s", sum(profiling.values()))
        for stage, elapsed in sorted(profiling.items(), key=lambda x: -x[1]):
            pct = (elapsed / total_time * 100) if total_time > 0 else 0
            md.append(f"| {stage} | {elapsed:.1f}s | {pct:.0f}% |")

    md.append("\n## Kesimpulan\n")
    md.append(f"Berdasarkan evaluasi terhadap {total} query, sistem GraphRAG & Agentic AI HERPA mencapai overall score **{overall.get('score', 0):.2f}/100** dengan grade **{overall.get('grade', 'N/A')}**.\n")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    return str(path)


def save_excel(eval_results: dict[str, Any]) -> str:
    _ensure_dir()
    path = OUTPUT_DIR / "evaluation_report.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return save_csv(eval_results) + " (openpyxl not installed)"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    agg = eval_results.get("aggregated_metrics", {})
    overall = agg.get("overall_score", {})
    ws["A1"] = "HERPA GraphRAG Evaluation"; ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Overall Score"; ws["B3"] = overall.get("score", 0)
    ws["A4"] = "Grade"; ws["B4"] = overall.get("grade", "N/A")
    ws["A5"] = "Total Queries"; ws["B5"] = eval_results.get("total_queries", 0)

    row = 7
    for section, data in [("Retrieval", agg.get("retrieval", {})), ("Judge", agg.get("judge", {})), ("Citation", agg.get("citation", {})), ("Graph", agg.get("graph", {}))]:
        ws[f"A{row}"] = section; ws[f"A{row}"].font = Font(bold=True); row += 1
        for k, v in data.items():
            ws[f"A{row}"] = k; ws[f"B{row}"] = v; row += 1
        row += 1

    ws2 = wb.create_sheet("Per-Query")
    per_query = eval_results.get("per_query_results", [])
    if per_query:
        headers = ["ID", "Query", "Category", "Precision@5", "Recall@5", "nDCG", "MRR", "Answer Relevancy", "Citation", "Graph Accuracy", "Latency (ms)", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
        for ri, r in enumerate(per_query, 2):
            ws2.cell(row=ri, column=1, value=r.get("id"))
            ws2.cell(row=ri, column=2, value=r.get("query", "")[:200])
            ws2.cell(row=ri, column=3, value=r.get("category"))
            ret = r.get("retrieval_metrics", {})
            ws2.cell(row=ri, column=4, value=ret.get("precision@5", 0))
            ws2.cell(row=ri, column=5, value=ret.get("recall@5", 0))
            ws2.cell(row=ri, column=6, value=ret.get("ndcg", 0))
            ws2.cell(row=ri, column=7, value=ret.get("mrr", 0))
            ws2.cell(row=ri, column=8, value=_safe_float(r.get("answer_relevancy")))
            ws2.cell(row=ri, column=9, value=r.get("citation_metrics", {}).get("accuracy", 0))
            ws2.cell(row=ri, column=10, value=r.get("graph_accuracy", {}).get("overall_accuracy", 0))
            ws2.cell(row=ri, column=11, value=r.get("latency", {}).get("total_ms", 0))
            ws2.cell(row=ri, column=12, value=r.get("status", ""))

    wb.save(path)
    return str(path)


def save_html(eval_results: dict[str, Any]) -> str:
    """Save HTML report."""
    from evaluation.export_html import generate_html_report
    _ensure_dir()
    path = str(OUTPUT_DIR / "evaluation_report.html")
    analytics = eval_results.get("analytics", {})
    return generate_html_report(eval_results, analytics, path)


def save_charts(eval_results: dict[str, Any]) -> list[str]:
    """Generate PNG charts if matplotlib available."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return []

    _ensure_dir()
    charts_dir = OUTPUT_DIR / "charts"
    charts_dir.mkdir(parents=True, exist_ok=True)
    saved = []

    agg = eval_results.get("aggregated_metrics", {})
    retrieval = agg.get("retrieval", {})
    judge = agg.get("judge", {})
    latency = agg.get("latency", {})

    # Precision@k and Recall@k
    fig, ax = plt.subplots(figsize=(8, 5))
    ks = [1, 3, 5]
    prec = [retrieval.get(f"precision@{k}", 0) for k in ks]
    rec = [retrieval.get(f"recall@{k}", 0) for k in ks]
    x = range(len(ks))
    ax.bar([i - 0.15 for i in x], prec, 0.3, label="Precision@k", color="#4472C4")
    ax.bar([i + 0.15 for i in x], rec, 0.3, label="Recall@k", color="#ED7D31")
    ax.set_xlabel("k"); ax.set_ylabel("Score"); ax.set_title("Precision@k vs Recall@k")
    ax.set_xticks(x); ax.set_xticklabels(ks); ax.legend(); ax.set_ylim(0, 1.1)
    p = charts_dir / "precision_recall.png"
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig); saved.append(str(p))

    # Answer Relevancy (single bar)
    fig, ax = plt.subplots(figsize=(6, 5))
    ans_rel = _safe_float(judge.get("answer_relevancy"))
    ax.bar(["Answer\nRelevancy"], [ans_rel], color="#4472C4")
    ax.set_ylabel("Score"); ax.set_title("LLM Evaluation: Answer Relevancy"); ax.set_ylim(0, 1.1)
    ax.text(0, ans_rel + 0.02, f"{ans_rel:.2f}", ha="center", fontsize=12)
    p = charts_dir / "answer_relevancy.png"
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig); saved.append(str(p))

    # Latency
    fig, ax = plt.subplots(figsize=(8, 5))
    stages = ["embedding", "neo4j", "retrieval", "llm", "judge", "total"]
    avgs = [latency.get(s, {}).get("avg", 0) for s in stages]
    p95s = [latency.get(s, {}).get("p95", 0) for s in stages]
    x = range(len(stages))
    ax.bar([i - 0.15 for i in x], avgs, 0.3, label="Avg", color="#4472C4")
    ax.bar([i + 0.15 for i in x], p95s, 0.3, label="P95", color="#ED7D31")
    ax.set_xlabel("Stage"); ax.set_ylabel("ms"); ax.set_title("Latency by Stage")
    ax.set_xticks(x); ax.set_xticklabels([s.capitalize() for s in stages]); ax.legend()
    p = charts_dir / "latency.png"
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig); saved.append(str(p))

    # Overall pie
    fig, ax = plt.subplots(figsize=(6, 6))
    components = agg.get("overall_score", {}).get("component_scores", {})
    if components:
        labels = [k.replace("_", "\n") for k in components.keys()]
        vals = list(components.values())
        colors = ["#4472C4", "#ED7D31", "#A5A5A5", "#FFC000", "#5B9BD5"]
        ax.pie(vals, labels=labels, autopct="%1.0f%%", colors=colors[:len(vals)])
        ax.set_title(f"Overall Score: {agg.get('overall_score', {}).get('score', 0):.1f}/100")
    p = charts_dir / "overall.png"
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig); saved.append(str(p))

    return saved


def print_error_report(eval_results: dict[str, Any]):
    """Print detailed error report for failed/skipped queries."""
    errors = eval_results.get("errors", [])
    if not errors:
        return

    try:
        from rich.console import Console
        from rich.table import Table
        from rich import box
        console = Console()

        console.print()
        t = Table(title=f"Error Report ({len(errors)} issues)", box=box.ROUNDED, show_lines=True)
        t.add_column("ID", style="bold")
        t.add_column("Query", max_width=40)
        t.add_column("Status")
        t.add_column("Errors")

        for r in errors:
            status = r.get("status", "FAIL")
            status_style = "[red]FAIL[/]" if status == "FAIL" else "[yellow]SKIP[/]"
            errs = r.get("errors", ["unknown"])
            main_err = errs[0] if errs else "unknown"
            if len(main_err) > 100:
                main_err = main_err[:100] + "..."
            t.add_row(str(r.get("id", "?")), r.get("query", "")[:40], status_style, main_err)

        console.print(t)
        console.print()

    except ImportError:
        print(f"\n  Errors ({len(errors)}):")
        for r in errors:
            print(f"    Q{r.get('id')}: {r.get('status')} — {'; '.join(r.get('errors', []))}")


def print_per_query_detail(results: list[dict[str, Any]], max_show: int = 10):
    """Print per-query detail with nodes, relations, context size, answer length."""
    try:
        from rich.console import Console
        from rich.table import Table
        from rich import box
        console = Console()

        t = Table(title=f"Per-Query Detail (first {min(max_show, len(results))})", box=box.ROUNDED, show_lines=True)
        t.add_column("ID", style="bold")
        t.add_column("Query", max_width=35)
        t.add_column("Nodes", justify="right")
        t.add_column("Rels", justify="right")
        t.add_column("Ctx", justify="right")
        t.add_column("Ans", justify="right")
        t.add_column("AnsRel", justify="right")
        t.add_column("Latency", justify="right")
        t.add_column("Status")

        for r in results[:max_show]:
            ans_rel = _safe_float(r.get("answer_relevancy"))
            lat = r.get("latency", {})
            status = r.get("status", "UNKNOWN")
            if status == "PASS":
                status_str = "[green]PASS[/]"
            elif status == "SKIP":
                status_str = "[yellow]SKIP[/]"
            else:
                status_str = "[red]FAIL[/]"
            cache = " [dim][C][/]" if r.get("_from_cache") else ""

            t.add_row(
                str(r.get("id", "?")),
                r.get("query", "")[:35] + cache,
                str(len(r.get("retrieved_nodes", []))),
                str(len(r.get("retrieved_relations", []))),
                str(r.get("context_size", 0)),
                str(r.get("answer_length", 0)),
                f"{ans_rel:.2f}",
                format_latency_ms(lat.get("total_ms", 0)),
                status_str,
            )
        console.print(t)
        if len(results) > max_show:
            console.print(f"  ... and {len(results) - max_show} more queries\n")

        # Print per-query diagnostics
        if results:
            t_diag = Table(title="Per-Query Diagnostics (first query)", box=box.SIMPLE_HEAVY, show_lines=False)
            t_diag.add_column("Key", style="bold")
            t_diag.add_column("Value", justify="right")
            first = results[0]
            t_diag.add_row("Retrieved Documents", str(first.get("retrieved_documents", 0)))
            t_diag.add_row("Retrieved Chunks", str(first.get("retrieved_chunks", 0)))
            t_diag.add_row("Retrieved Text", f"{first.get('retrieved_text_chars', 0)} chars")
            t_diag.add_row("Context Builder", first.get("context_builder_status", "N/A"))
            t_diag.add_row("Prompt Builder", first.get("prompt_builder_status", "N/A"))
            t_diag.add_row("Judge Ready", first.get("judge_ready", "N/A"))
            contexts = first.get("contexts", [])
            if contexts:
                preview = contexts[0][:150] + "..." if len(contexts[0]) > 150 else contexts[0]
                t_diag.add_row("Context Preview", preview)
            console.print(t_diag)

            audit = first.get("retrieval_audit", {})
            if audit and audit.get("entity_audit"):
                ea = audit["entity_audit"]
                console.print("[bold]Retrieval Audit (Q1):[/]")
                console.print(f"  Detected Entities: {ea.get('detected_entities', [])}")
                console.print(f"  Detected Diseases: {ea.get('detected_diseases', [])}")
                console.print(f"  Detected Compounds: {ea.get('detected_compounds', [])}")
                console.print(f"  Missing Entity: {ea.get('missing_entity', False)}")
                console.print(f"  Fallback Used: {ea.get('fallback_used', False)}")
                if audit.get("warnings"):
                    console.print(f"  Warnings: {audit['warnings']}")
                console.print()

    except ImportError:
        for r in results[:max_show]:
            lat = r.get("latency", {})
            print(f"  Q{r.get('id')}: Nodes={len(r.get('retrieved_nodes',[]))} "
                  f"AnsRel={_safe_float(r.get('answer_relevancy')):.2f} "
                  f"Lat={format_latency_ms(lat.get('total_ms',0))} "
                  f"Status={r.get('status','?')}")
