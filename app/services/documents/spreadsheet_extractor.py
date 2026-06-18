import csv
from io import BytesIO

from openpyxl import load_workbook

from app.core.config import Settings


def extract_xlsx(data: bytes, settings: Settings) -> dict:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    chunks: list[str] = []
    sheets: list[dict] = []
    for worksheet in workbook.worksheets[:20]:
        rows: list[str] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row_index > 1000:
                break
            values = ["" if value is None else str(value) for value in row[:50]]
            rows.append(" | ".join(values))
        section = f"[Sheet: {worksheet.title}]\n" + "\n".join(rows)
        chunks.append(section)
        sheets.append({"name": worksheet.title, "rows": len(rows)})
    text = "\n\n".join(chunks)
    return {
        "text": text[: settings.max_document_characters],
        "sheets": sheets,
        "truncated": len(text) > settings.max_document_characters,
        "needs_vision": False,
    }


def extract_csv(data: bytes, settings: Settings) -> dict:
    decoded = data.decode("utf-8-sig", errors="replace")
    rows: list[str] = []
    for index, row in enumerate(csv.reader(decoded.splitlines())):
        if index >= 2000:
            break
        rows.append(" | ".join(row[:50]))
    text = "\n".join(rows)
    return {
        "text": text[: settings.max_document_characters],
        "truncated": len(text) > settings.max_document_characters,
        "needs_vision": False,
    }
